"""Consolida los welding books individuales en uno solo de proyecto.

Para una carpeta con PDFs ya procesados:
  1. Lee cada `*_welding_book.csv` (uno por PDF marcado).
  2. Extrae el cajetín del PDF correspondiente (o lo lee de `cajetines.csv`
     si existe, para no re-parsearlo).
  3. Genera un único CSV/XLSX con TODAS las costuras + metadatos de la
     línea/documento de cada una.

Uso:
    .venv/bin/python -m tools.consolidar_welding_book \\
        --carpeta revision_10pdfs \\
        --salida welding_book_consolidado.xlsx
"""

from __future__ import annotations

import argparse
import csv
import sys
from pathlib import Path

_HERE = Path(__file__).resolve().parent
if str(_HERE.parent) not in sys.path:
    sys.path.insert(0, str(_HERE.parent))

from tools.extraer_cajetin import DatosCajetin, extraer_cajetin  # noqa: E402


COLUMNAS_SALIDA = [
    # Metadatos del isométrico
    "linea", "documento", "sheet", "revision", "estado", "diametro",
    "cliente", "proyecto", "archivo_pdf",
    # Datos de la costura (del welding_book individual)
    "n_costura_local", "etiqueta", "tipo", "x", "y", "radio",
    "distancia_tuberia", "id_elemento",
    "itemcode", "diametro_costura", "inspeccionado", "observaciones",
]

# Mapeo desde las cabeceras en español del welding_book individual.
MAPEO_WB = {
    "Nº costura": "n_costura_local",
    "Etiqueta": "etiqueta",
    "Tipo": "tipo",
    "X": "x",
    "Y": "y",
    "Radio detectado": "radio",
    "Distancia a tubería": "distancia_tuberia",
    "ID elemento origen": "id_elemento",
    "Itemcode": "itemcode",
    "Diámetro": "diametro_costura",
    "Inspeccionado": "inspeccionado",
    "Observaciones": "observaciones",
}


def cargar_cajetines_cache(csv_path: Path) -> dict[str, DatosCajetin]:
    """Lee `cajetines.csv` y devuelve dict {archivo_pdf: DatosCajetin}."""
    if not csv_path.is_file():
        return {}
    out: dict[str, DatosCajetin] = {}
    with open(csv_path, encoding="utf-8-sig", newline="") as f:
        for fila in csv.DictReader(f, delimiter=";"):
            d = DatosCajetin(
                archivo=fila.get("archivo", ""),
                linea=fila.get("linea", ""),
                documento=fila.get("documento", ""),
                sheet=fila.get("sheet", ""),
                revision=fila.get("revision", ""),
                estado=fila.get("estado", ""),
                diametro=fila.get("diametro", ""),
                cliente=fila.get("cliente", ""),
                proyecto=fila.get("proyecto", ""),
                notas=[fila["notas"]] if fila.get("notas") else [],
            )
            out[d.archivo] = d
    return out


def localizar_pdf(carpeta: Path, wb_csv: Path) -> Path | None:
    """Dado un `*_welding_book.csv`, devuelve el PDF original."""
    nombre = wb_csv.name
    if not nombre.endswith("_welding_book.csv"):
        return None
    raiz = nombre[: -len("_welding_book.csv")]
    pdf = carpeta / f"{raiz}.pdf"
    return pdf if pdf.is_file() else None


def cargar_welding_book(wb_csv: Path) -> list[dict]:
    out: list[dict] = []
    with open(wb_csv, encoding="utf-8-sig", newline="") as f:
        for fila in csv.DictReader(f, delimiter=";"):
            mapeado = {MAPEO_WB.get(k, k): v for k, v in fila.items()}
            out.append(mapeado)
    return out


def consolidar(carpeta: Path) -> list[dict]:
    cache_cajetines = cargar_cajetines_cache(carpeta / "cajetines.csv")
    if cache_cajetines:
        print(f"  cajetines.csv encontrado, {len(cache_cajetines)} entradas reutilizadas.")

    wbs = sorted(carpeta.glob("*_welding_book.csv"))
    print(f"  {len(wbs)} welding books individuales encontrados.")

    filas: list[dict] = []
    for wb in wbs:
        pdf = localizar_pdf(carpeta, wb)
        if pdf is None:
            print(f"    SKIP {wb.name}: PDF no encontrado")
            continue

        cajetin = cache_cajetines.get(pdf.name)
        if cajetin is None:
            print(f"    Extrayendo cajetín en vivo: {pdf.name}")
            cajetin = extraer_cajetin(pdf)

        notas = " | ".join(cajetin.notas) if cajetin.notas else ""

        for fila_wb in cargar_welding_book(wb):
            fila = {
                "linea": cajetin.linea,
                "documento": cajetin.documento,
                "sheet": cajetin.sheet,
                "revision": cajetin.revision,
                "estado": cajetin.estado,
                "diametro": cajetin.diametro,
                "cliente": cajetin.cliente,
                "proyecto": cajetin.proyecto,
                "archivo_pdf": pdf.name,
            }
            fila.update({c: fila_wb.get(c, "") for c in COLUMNAS_SALIDA
                         if c not in fila})
            fila["observaciones"] = (
                (fila.get("observaciones", "") or "")
                + ((" | " + notas) if notas and not fila.get("observaciones") else "")
            ).strip(" |")
            filas.append(fila)
    return filas


def escribir_csv(filas: list[dict], salida: Path) -> None:
    with open(salida, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=COLUMNAS_SALIDA, delimiter=";")
        w.writeheader()
        for r in filas:
            w.writerow({c: r.get(c, "") for c in COLUMNAS_SALIDA})


def escribir_xlsx(filas: list[dict], salida: Path) -> bool:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        return False
    wb = Workbook()
    ws = wb.active
    ws.title = "Welding Book"
    ws.append(COLUMNAS_SALIDA)
    cabecera = Font(bold=True, color="FFFFFF")
    relleno = PatternFill("solid", fgColor="2F5496")
    for c in ws[1]:
        c.font = cabecera
        c.fill = relleno
        c.alignment = Alignment(horizontal="center")
    for r in filas:
        ws.append([r.get(c, "") for c in COLUMNAS_SALIDA])
    # Ajustar anchos básicos
    anchos = {
        "linea": 22, "documento": 22, "sheet": 18, "revision": 8,
        "estado": 24, "diametro": 9, "cliente": 12, "proyecto": 30,
        "archivo_pdf": 50, "n_costura_local": 8, "etiqueta": 10, "tipo": 6,
        "x": 9, "y": 9, "radio": 8, "distancia_tuberia": 9, "id_elemento": 14,
        "itemcode": 14, "diametro_costura": 9, "inspeccionado": 12,
        "observaciones": 40,
    }
    for i, col in enumerate(COLUMNAS_SALIDA, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = anchos.get(col, 12)
    ws.freeze_panes = "A2"
    wb.save(salida)
    return True


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--carpeta", required=True,
                    help="Carpeta con los *_welding_book.csv y *.pdf")
    ap.add_argument("--salida", default="welding_book_consolidado.xlsx",
                    help="Ruta del fichero de salida (.csv o .xlsx)")
    args = ap.parse_args()

    carpeta = Path(args.carpeta)
    salida = Path(args.salida)
    if not carpeta.is_dir():
        sys.exit(f"No existe la carpeta: {carpeta}")

    print(f"Consolidando welding books de {carpeta}…")
    filas = consolidar(carpeta)

    if not filas:
        sys.exit("Sin costuras para consolidar.")

    if salida.suffix.lower() == ".xlsx":
        if not escribir_xlsx(filas, salida):
            print("openpyxl no disponible, fallback a CSV.")
            salida = salida.with_suffix(".csv")
            escribir_csv(filas, salida)
    else:
        escribir_csv(filas, salida)

    print(f"\nTotal costuras consolidadas: {len(filas)}")
    print(f"PDFs únicos: {len(set(r['archivo_pdf'] for r in filas))}")
    print(f"Líneas únicas: {len(set(r['linea'] for r in filas if r['linea']))}")
    print(f"Salida: {salida}")


if __name__ == "__main__":
    main()
