"""Enriquece un welding book con datos del BOM del propio PDF.

Para 1 PDF:
  1. Carga las costuras detectadas (vía detector del addon).
  2. Parsea el BOM (`extraer_bom`).
  3. Detecta piece marks dibujados sobre el plano: textos numéricos de
     1-2 dígitos en la zona del dibujo cuyo valor coincide con un PT NO
     del BOM.
  4. Para cada costura, encuentra los 2 piece marks más cercanos
     (dentro de un radio configurable) — son los componentes que se
     sueldan en esa costura.
  5. Hereda al welding book: itemcode, descripción, diámetro de cada
     componente, lista de piece marks adyacentes.

Salida: <pdf>_welding_book_enriquecido.xlsx (CSV de fallback).

Uso:
    .venv/bin/python -m tools.enriquecer_welding_book <pdf>
"""

from __future__ import annotations

import argparse
import csv
import sys
from dataclasses import dataclass
from math import hypot
from pathlib import Path

_HERE = Path(__file__).resolve().parent
if str(_HERE.parent) not in sys.path:
    sys.path.insert(0, str(_HERE.parent))

# Inkex / addon
_INK_EXT = "/usr/share/inkscape/extensions"
_USR_EXT = str(Path.home() / ".config/inkscape/extensions")
for p in (_INK_EXT, _USR_EXT):
    if p not in sys.path:
        sys.path.insert(0, p)

import inkex  # noqa: E402

from tools.extraer_bom import ItemBOM, extraer_bom  # noqa: E402
from tools.extraer_cajetin import extraer_cajetin  # noqa: E402
from tools.pdf_extractor import extraer_pdf  # noqa: E402
from tramear_core.seam_detector import detectar_costuras  # noqa: E402


RADIO_PM_DEF = 70.0   # pt — alcance para considerar piece mark de una costura
RADIO_RESCATE = 35.0  # pt — radio para buscar costuras perdidas alrededor de PM huérfano
ZONA_X_MAX = 770
ZONA_Y_MIN = 50
ZONA_Y_MAX = 600

# Categorías cuyos componentes SE SUELDAN a la tubería (cada PM esperado
# tiene típicamente ≥1 costura cerca). El resto (PIPE SUPPORTS, BOLTS,
# GASKETS) son atornillados o soldados a estructura ajena al circuito.
CATEGORIAS_SOLDABLES = {"PIPE", "FITTINGS", "FLANGES",
                         "VALVES", "INSTRUMENTS", "SPECIALS"}


@dataclass
class PieceMark:
    pt_no: str
    x: float
    y: float


def detectar_piece_marks(pdf: Path, items_bom: list[ItemBOM]) -> list[PieceMark]:
    pts_validos = {i.pt_no for i in items_bom if i.pt_no}
    extr = extraer_pdf(pdf, paginas=[0])
    if not extr.paginas:
        return []
    pag = extr.paginas[0]
    out: list[PieceMark] = []
    for t in pag.textos:
        s = t.texto.strip()
        if not s.isdigit() or not (1 <= int(s) <= 99) or len(s) > 2:
            continue
        if not (t.x < ZONA_X_MAX and ZONA_Y_MIN < t.y < ZONA_Y_MAX):
            continue
        if s not in pts_validos:
            # Texto numérico que no coincide con ningún PT — probable cota.
            continue
        out.append(PieceMark(pt_no=s, x=t.x, y=t.y))
    return out


def cargar_svg(pdf: Path) -> Path:
    cache = Path(_HERE) / "_artifacts" / "validacion" / "cache_svg"
    cache.mkdir(parents=True, exist_ok=True)
    svg = cache / f"{pdf.stem}.svg"
    if not svg.is_file():
        import subprocess
        subprocess.run(
            ["inkscape", "--pdf-poppler", "--export-type=svg",
             f"--export-filename={svg}", str(pdf)],
            check=True, capture_output=True,
        )
    return svg


def costuras_de(svg: Path):
    arbol = inkex.load_svg(str(svg)).getroot()
    return detectar_costuras(
        arbol, radio_min=1.4, radio_max=2.5, tolerancia_linea=1.5,
        solo_sobre_linea=True, detectar_field_welds=False,
        filtrar_puntas_flecha=True,
    )


def candidatos_brutos(svg: Path):
    """Devuelve los círculos candidatos brutos (rellenos, sobre línea),
    SIN aplicar filtros agresivos de triángulo de flecha. Sirve para la
    pasada de rescate guiada por BOM."""
    from tramear_core.seam_detector import (
        _buscar_circulos, _deduplicar, _distancia_minima_a_segmentos,
        _extraer_segmentos_lineales,
    )
    arbol = inkex.load_svg(str(svg)).getroot()
    crudos = _buscar_circulos(arbol, 1.4, 2.5)
    segs = _extraer_segmentos_lineales(arbol)
    en_linea = []
    for c in crudos:
        d = _distancia_minima_a_segmentos(c.x, c.y, segs)
        c.distancia_linea = d
        c.sobre_linea = d <= 1.5
        if c.sobre_linea and c.relleno:
            en_linea.append(c)
    return _deduplicar(en_linea)


def rescatar_por_bom(costuras, candidatos, piece_marks: list[PieceMark],
                     items_bom: list[ItemBOM],
                     radio_rescate: float = RADIO_RESCATE,
                     dist_dup: float = 1.5) -> list:
    """Recupera costuras perdidas usando el BOM como ground truth.

    Para cada piece mark de categoría soldable que NO tiene costura
    detectada en su entorno, busca en `candidatos` (círculos brutos
    rellenos sobre línea, antes de filtros agresivos) algún punto en
    radio `radio_rescate` y lo añade como costura.

    Devuelve la lista combinada (originales + rescatadas).
    """
    by_pt = {i.pt_no: i for i in items_bom}
    pms_soldables = [
        pm for pm in piece_marks
        if (it := by_pt.get(pm.pt_no)) and it.categoria in CATEGORIAS_SOLDABLES
    ]

    rescatadas = []
    ya = list(costuras)
    for pm in pms_soldables:
        # ¿Hay alguna costura ya detectada cerca?
        tiene_cerca = any(
            hypot(c.x - pm.x, c.y - pm.y) <= radio_rescate
            for c in ya + rescatadas
        )
        if tiene_cerca:
            continue
        # Buscar el candidato bruto más cercano dentro del radio.
        mejor = None
        mejor_d = float("inf")
        for cand in candidatos:
            if any(hypot(cand.x - o.x, cand.y - o.y) < dist_dup
                   for o in ya + rescatadas):
                continue
            d = hypot(cand.x - pm.x, cand.y - pm.y)
            if d <= radio_rescate and d < mejor_d:
                mejor = cand
                mejor_d = d
        if mejor is not None:
            rescatadas.append(mejor)

    return list(costuras) + rescatadas, rescatadas


def asociar(costuras, piece_marks: list[PieceMark],
            items_bom: list[ItemBOM], radio: float):
    """Para cada costura devuelve (pm_a, pm_b, item_a, item_b)."""
    by_pt = {i.pt_no: i for i in items_bom}
    out = []
    for cs in costuras:
        candidatos = [
            (hypot(pm.x - cs.x, pm.y - cs.y), pm)
            for pm in piece_marks
        ]
        candidatos.sort(key=lambda r: r[0])
        elegidos = [pm for d, pm in candidatos if d <= radio][:2]
        # Si solo hay 1 cercano, dejamos el segundo vacío.
        a = elegidos[0] if elegidos else None
        b = elegidos[1] if len(elegidos) > 1 else None
        out.append((cs, a, b, by_pt.get(a.pt_no) if a else None,
                    by_pt.get(b.pt_no) if b else None))
    return out


COLUMNAS = [
    "isometrico", "n_costura", "etiqueta", "diametro",
    "pm_1", "itemcode_1", "descripcion_1",
    "pm_2", "itemcode_2", "descripcion_2",
    "x", "y", "radio",
]


def construir_filas(asoc, isometrico: str) -> list[dict]:
    filas = []
    for i, (cs, pm_a, pm_b, it_a, it_b) in enumerate(asoc, start=1):
        diam_costura = _diam_costura(it_a, it_b)
        filas.append({
            "isometrico": isometrico,
            "n_costura": i,
            "etiqueta": f"W-{i:03d}",
            "diametro": diam_costura,
            "pm_1": pm_a.pt_no if pm_a else "",
            "itemcode_1": it_a.itemcode if it_a else "",
            "descripcion_1": it_a.descripcion if it_a else "",
            "pm_2": pm_b.pt_no if pm_b else "",
            "itemcode_2": it_b.itemcode if it_b else "",
            "descripcion_2": it_b.descripcion if it_b else "",
            "x": round(cs.x, 2),
            "y": round(cs.y, 2),
            "radio": round(cs.radio, 2),
        })
    return filas


def _diam_costura(it_a, it_b) -> str:
    diams = [d for d in (it_a.diametro if it_a else "",
                          it_b.diametro if it_b else "") if d]
    if not diams:
        return ""
    if len(set(diams)) == 1:
        return diams[0]
    # Si difieren (reducción), devolver el mayor de los primeros números.
    def primer_num(s: str) -> float:
        import re
        m = re.match(r"\s*(\d+(?:\.\d+)?)", s)
        return float(m.group(1)) if m else 0.0
    return max(diams, key=primer_num)


def escribir_xlsx(filas, salida: Path) -> bool:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        return False
    wb = Workbook()
    ws = wb.active
    ws.title = "Welding Book"
    ws.append(COLUMNAS)
    cab = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="2F5496")
    for c in ws[1]:
        c.font = cab
        c.fill = fill
        c.alignment = Alignment(horizontal="center")
    for r in filas:
        ws.append([r.get(c, "") for c in COLUMNAS])
    anchos = {"isometrico": 22, "descripcion_1": 50, "descripcion_2": 50,
              "etiqueta": 10, "n_costura": 6, "radio": 7,
              "x": 9, "y": 9, "pm_1": 6, "pm_2": 6,
              "itemcode_1": 12, "itemcode_2": 12, "diametro": 10}
    for i, col in enumerate(COLUMNAS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = \
            anchos.get(col, 12)
    ws.freeze_panes = "A2"
    wb.save(salida)
    return True


def escribir_csv(filas, salida: Path) -> None:
    with open(salida, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=COLUMNAS, delimiter=";")
        w.writeheader()
        w.writerows(filas)


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("pdf")
    ap.add_argument("--radio", type=float, default=RADIO_PM_DEF,
                    help="Radio de búsqueda de piece marks (pt).")
    ap.add_argument("--salida", default=None)
    args = ap.parse_args()

    pdf = Path(args.pdf)
    if not pdf.is_file():
        sys.exit(f"PDF no existe: {pdf}")

    print(f"Procesando: {pdf.name}")
    svg = cargar_svg(pdf)
    costuras = costuras_de(svg)
    print(f"  costuras detectadas: {len(costuras)}")

    items = extraer_bom(pdf)
    print(f"  items BOM: {len(items)}")

    pms = detectar_piece_marks(pdf, items)
    print(f"  piece marks en plano: {len(pms)} "
          f"({sorted(set(p.pt_no for p in pms))})")

    cajetin = extraer_cajetin(pdf)
    isometrico = cajetin.sheet or cajetin.linea or pdf.stem

    # Pasada de rescate guiada por BOM: si un PM soldable no tiene costura
    # cerca, recuperar el círculo bruto más próximo (relajando el filtro
    # de triángulo que descartó).
    crudos = candidatos_brutos(svg)
    costuras, rescatadas = rescatar_por_bom(costuras, crudos, pms, items)
    if rescatadas:
        print(f"  rescatadas guiadas por BOM: {len(rescatadas)}")
        for r in rescatadas:
            print(f"    + ({r.x:.1f}, {r.y:.1f}) r={r.radio:.2f}")
    print(f"  costuras totales: {len(costuras)}")

    asoc = asociar(costuras, pms, items, args.radio)
    filas = construir_filas(asoc, isometrico)
    cob = sum(1 for f in filas if f["pm_1"]) / max(len(filas), 1)
    cob2 = sum(1 for f in filas if f["pm_1"] and f["pm_2"]) / max(len(filas), 1)
    print(f"  cobertura ≥1 piece mark: {cob*100:.0f}%")
    print(f"  cobertura ≥2 piece marks: {cob2*100:.0f}%")

    salida = (Path(args.salida) if args.salida
              else pdf.parent / f"{pdf.stem}_welding_book_enriquecido.xlsx")
    if salida.suffix.lower() == ".xlsx":
        if not escribir_xlsx(filas, salida):
            salida = salida.with_suffix(".csv")
            escribir_csv(filas, salida)
    else:
        escribir_csv(filas, salida)
    print(f"  salida: {salida}")


if __name__ == "__main__":
    main()
