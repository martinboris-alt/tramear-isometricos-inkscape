"""Exportador del welding book a CSV y XLSX."""

from __future__ import annotations

import csv
from datetime import datetime
from pathlib import Path
from typing import Sequence

from .seam_detector import Costura


COLUMNAS = (
    "Nº costura",
    "Etiqueta",
    "Tipo",            # SW = shop weld, FW = field weld
    "X",
    "Y",
    "Radio detectado",
    "Distancia a tubería",
    "ID elemento origen",
    "Itemcode",        # reservado para v2 (identificación de materiales)
    "Diámetro",        # reservado para v2
    "Inspeccionado",   # casilla para que el operario rellene
    "Observaciones",
)


def exportar_csv(
    numeradas: Sequence[tuple[int, Costura]],
    ruta_csv: Path,
    prefijo: str = "W-",
) -> Path:
    ruta_csv = Path(ruta_csv)
    ruta_csv.parent.mkdir(parents=True, exist_ok=True)
    with ruta_csv.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(COLUMNAS)
        for n, c in numeradas:
            w.writerow([
                n,
                f"{prefijo}{n:03d}",
                c.tipo,
                f"{c.x:.3f}",
                f"{c.y:.3f}",
                f"{c.radio:.3f}",
                f"{c.distancia_linea:.3f}",
                c.fuente_id,
                "", "", "", "",
            ])
    return ruta_csv


def exportar_xlsx(
    numeradas: Sequence[tuple[int, Costura]],
    ruta_xlsx: Path,
    prefijo: str = "W-",
    nombre_iso: str = "",
) -> Path | None:
    """Exporta a XLSX si openpyxl está disponible. Si no, devuelve None."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        return None

    ruta_xlsx = Path(ruta_xlsx)
    ruta_xlsx.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Welding Book"

    ws["A1"] = "WELDING BOOK"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Isométrico: {nombre_iso or '(sin nombre)'}"
    ws["A3"] = f"Generado: {datetime.now():%Y-%m-%d %H:%M}"
    ws["A4"] = f"Total costuras: {len(numeradas)}"

    fila_cab = 6
    cab_font = Font(bold=True, color="FFFFFF")
    cab_fill = PatternFill("solid", fgColor="2E5F8C")
    for j, col in enumerate(COLUMNAS, start=1):
        celda = ws.cell(row=fila_cab, column=j, value=col)
        celda.font = cab_font
        celda.fill = cab_fill
        celda.alignment = Alignment(horizontal="center")

    for i, (n, c) in enumerate(numeradas, start=fila_cab + 1):
        ws.cell(row=i, column=1, value=n)
        ws.cell(row=i, column=2, value=f"{prefijo}{n:03d}")
        ws.cell(row=i, column=3, value=c.tipo)
        ws.cell(row=i, column=4, value=round(c.x, 3))
        ws.cell(row=i, column=5, value=round(c.y, 3))
        ws.cell(row=i, column=6, value=round(c.radio, 3))
        ws.cell(row=i, column=7, value=round(c.distancia_linea, 3))
        ws.cell(row=i, column=8, value=c.fuente_id)

    # Anchos razonables
    for col_letra, ancho in zip("ABCDEFGHIJKL",
                                 (10, 12, 7, 10, 10, 14, 18, 22, 14, 12, 14, 30)):
        ws.column_dimensions[col_letra].width = ancho

    wb.save(ruta_xlsx)
    return ruta_xlsx
